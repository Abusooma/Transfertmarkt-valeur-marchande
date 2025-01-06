from loguru import logger
import asyncio
from datetime import datetime
import time
import sys
import threading
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from players import ScraperTransferMarkt

# Configuration du logger
logger.remove()
logger.add(sys.stdout, level="INFO", colorize=True,
           format="<green>{time:HH:mm:ss}</green> | <level>{message}</level>")
logger.add("logs/fichier.log", level="ERROR",
           format="{time:YYYY-MM-DD HH:mm:ss} | {level} | {message}", rotation="1 MB")


class RealTimeChronometre:
    def __init__(self):
        self.debut = time.time()
        self.arret = False
        self.thread = None

    def demarrer(self):
        def afficher_temps():
            while not self.arret:
                temps_ecoule = time.time() - self.debut
                minutes = int(temps_ecoule // 60)
                secondes = int(temps_ecoule % 60)
                sys.stdout.write(
                    f"\rTemps écoulé: {minutes:02d}:{secondes:02d}")
                sys.stdout.flush()
                time.sleep(1)

        self.thread = threading.Thread(target=afficher_temps, daemon=True)
        self.thread.start()

    def arreter(self):
        self.arret = True
        if self.thread:
            self.thread.join()
        temps_total = time.time() - self.debut
        return temps_total


class MiseAJourValeursJoueurs:
    def __init__(self, fichier_entree: str, fichier_sortie: str):
        self.fichier_entree = fichier_entree
        self.fichier_sortie = fichier_sortie
        self.scraper = ScraperTransferMarkt(max_threads=3)
        self.chronometre = RealTimeChronometre()

    def formater_nom2(self, nom_original: str) -> str:
        parties = nom_original.split()
        index_majuscule = next(
            (i for i, mot in enumerate(parties) if mot.isupper()), None)

        if index_majuscule is not None:
            mot_majuscule = parties[index_majuscule]
            autres_parties = parties[:index_majuscule] + \
                parties[index_majuscule+1:]
            return f"{mot_majuscule} {' '.join(autres_parties)}"

        return nom_original

    async def mettre_a_jour(self):
        logger.info("Début du Processus")
        self.chronometre.demarrer()

        df = pd.read_excel(self.fichier_entree, dtype={"NOM": str})
        noms_joueurs = df["NOM"].tolist()

        try:
            valeurs_joueurs = await asyncio.to_thread(
                self.scraper.recuperer_valeurs_joueurs,
                noms_joueurs
            )
        except Exception as e:
            logger.error(f"Erreur durant le scraping : {e}")
            self.chronometre.arreter()
            raise

        donnees_mises_a_jour = []
        date_courante = datetime.now().strftime("%d/%m/%Y")

        for _, ligne in df.iterrows():
            nom = ligne["NOM"]
            valeur_joueur = valeurs_joueurs.get(nom, None)

            if valeur_joueur and valeur_joueur.nom_transfermarkt:
                parties_nom = valeur_joueur.nom_transfermarkt.split()
                if len(parties_nom) >= 3:
                    nom2 = f"{parties_nom[-2].upper()} {parties_nom[-1].upper()} {' '.join(parties_nom[:-2])}"
                else:
                    nom2 = f"{parties_nom[-1].upper()} {' '.join(parties_nom[:-1])}"
            else:
                nom2 = self.formater_nom2(nom)

            donnees = {
                "NOM": valeur_joueur.nom_transfermarkt if valeur_joueur else "",
                "DOB": valeur_joueur.date_naissance if valeur_joueur and valeur_joueur.date_naissance else "",
                "DATE": date_courante,
                "VALEUR": valeur_joueur.valeur if valeur_joueur else 0.0,
                "NOM2": nom2,
                "FIN-CONTRAT": valeur_joueur.fin_contrat if valeur_joueur and valeur_joueur.fin_contrat else "",
                "CONTROLE": valeur_joueur.controle if valeur_joueur else ""
            }
            donnees_mises_a_jour.append(donnees)

        df_mise_a_jour = pd.DataFrame(donnees_mises_a_jour)

        colonnes = ["NOM", "DOB", "DATE", "VALEUR",
                    "NOM2", "FIN-CONTRAT", "CONTROLE"]
        for col in colonnes:
            if col not in df_mise_a_jour.columns:
                df_mise_a_jour[col] = ""

        df_mise_a_jour = df_mise_a_jour[colonnes]

        with pd.ExcelWriter(self.fichier_sortie, engine='openpyxl') as writer:
            df_mise_a_jour.to_excel(writer, index=False, sheet_name='Sheet1')

            workbook = writer.book
            worksheet = workbook['Sheet1']

            for column in worksheet.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column[0]
                                            .column_letter].width = adjusted_width

            yellow_fill = PatternFill(start_color="FFFF00",
                                      end_color="FFFF00",
                                      fill_type="solid")

            controle_col = None
            for idx, col in enumerate(worksheet[1], 1):
                if col.value == "CONTROLE":
                    controle_col = idx
                    break

            if controle_col:
                for row in worksheet.iter_rows(min_row=2):
                    if row[controle_col-1].value == "A verifier":
                        for cell in row:
                            cell.fill = yellow_fill

        temps_total = self.chronometre.arreter()
        logger.info(
            f"Processus terminé. Fichier enregistré sous {self.fichier_sortie}")
        logger.info(f"Temps total d'exécution : {temps_total:.2f} secondes")


async def main():
    fichier_entree = "pour-inverser-sortie1.xls"
    fichier_sortie = "resultat_essai.xlsx"

    mise_a_jour = MiseAJourValeursJoueurs(fichier_entree, fichier_sortie)
    try:
        await mise_a_jour.mettre_a_jour()
    except Exception as e:
        logger.error(f"Une erreur s'est produite : {e}")
    finally:
        mise_a_jour.scraper.cache.fermer()

if __name__ == "__main__":
    asyncio.run(main())
