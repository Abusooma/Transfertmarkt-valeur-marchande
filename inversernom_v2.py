import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os


class InverseurNoms:
   def __init__(self, fichier_entree, fichier_sortie):
       self.fichier_entree = fichier_entree
       self.fichier_sortie = fichier_sortie

   def inverser_nom(self, nom):
       if not isinstance(nom, str):
           return nom
       parties_nom = nom.strip().split()
       if len(parties_nom) >= 3:
           nom = f"{parties_nom[-2].upper()} {parties_nom[-1].upper()} {' '.join(parties_nom[:-2])}"
       else:
           nom = f"{parties_nom[-1].upper()} {' '.join(parties_nom[:-1])}"
       return nom

   def traiter_fichier(self):
       try:
           df = pd.read_excel(self.fichier_entree, engine='openpyxl')
       except:
           try:
               df = pd.read_excel(self.fichier_entree, engine='xlrd')
           except:
               raise ValueError("Impossible de lire le fichier Excel")

       if 'NOM' not in df.columns:
           raise ValueError("La colonne 'NOM' est introuvable")

       df_sortie = df.copy()
       nom_index = df_sortie.columns.get_loc('NOM')
       noms_inverses = df_sortie['NOM'].apply(self.inverser_nom)

       colonnes = list(df_sortie.columns)
       colonnes.insert(nom_index + 1, 'NOM_INVERSE')
       df_sortie = df_sortie.reindex(columns=colonnes)
       df_sortie['NOM_INVERSE'] = noms_inverses

       try:
           df_sortie.to_excel(self.fichier_sortie,
                              index=False, engine='openpyxl')
       except:
           try:
               df_sortie.to_excel(self.fichier_sortie,
                                  index=False, engine='xlwt')
           except:
               raise ValueError("Impossible de sauvegarder le fichier Excel")

       if self.fichier_sortie.endswith('.xlsx'):
           self.ajuster_largeur_colonnes(self.fichier_sortie)

       print(
           f"Fichier traité avec succès. Sauvegardé dans {self.fichier_sortie}")

   def ajuster_largeur_colonnes(self, fichier: str):
       wb = load_workbook(fichier)
       sheet = wb.active

       for col in sheet.columns:
           max_length = 0
           col_letter = get_column_letter(col[0].column)
           for cell in col:
               try:
                   if cell.value:
                       max_length = max(max_length, len(str(cell.value)))
               except:
                   pass
           adjusted_width = max_length + 2
           sheet.column_dimensions[col_letter].width = adjusted_width

       wb.save(fichier)


def main():
   fichier_entree = 'PourRUNTR8JANVIER.xls'
   fichier_sortie = 're.xlsx'

   inverseur = InverseurNoms(fichier_entree, fichier_sortie)
   inverseur.traiter_fichier()


if __name__ == "__main__":
   main()
