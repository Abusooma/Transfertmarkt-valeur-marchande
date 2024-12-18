import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os


class InverseurNoms:
    def __init__(self, fichier_entree, fichier_sortie):
        self.fichier_entree = fichier_entree
        self.fichier_sortie = fichier_sortie

    def inverser_nom(self, nom):
        if not isinstance(nom, str):
            return nom

        parties_nom = nom.strip().split()

        if len(parties_nom) >= 3:
            nom = f"{parties_nom[-2].upper()} {parties_nom[-1].upper()} {' '.join(parties_nom[:-2])}"
        else:
            nom = f"{parties_nom[-1].upper()} {' '.join(parties_nom[:-1])}"

        return nom

    def traiter_fichier(self):
        extension = os.path.splitext(self.fichier_entree)[1].lower()

        if extension == '.xls':
            df = pd.read_excel(self.fichier_entree, engine='xlrd')
        else:
            df = pd.read_excel(self.fichier_entree)

        if 'Entree' not in df.columns:
            raise ValueError(
                "La colonne 'Entree' est introuvable dans le fichier Excel")

        df_sortie = pd.DataFrame(
            {'NOM': df['Entree'].apply(self.inverser_nom)})

        extension_sortie = os.path.splitext(self.fichier_sortie)[1].lower()

        if extension_sortie == '.xls':
            df_sortie.to_excel(self.fichier_sortie, index=False, engine='xlwt')
        else:
            df_sortie.to_excel(self.fichier_sortie, index=False)

        if extension_sortie in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
            self.ajuster_largeur_colonnes(self.fichier_sortie)

        print(
            f"Fichier traité avec succès. Sauvegardé dans {self.fichier_sortie}")

    def ajuster_largeur_colonnes(self, fichier: str):
        wb = load_workbook(fichier)
        sheet = wb.active

        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = max_length + 2
            sheet.column_dimensions[col_letter].width = adjusted_width

        wb.save(fichier)


def main():
    fichier_entree = 'Fichier-transf2.xls'  # à remplacer par ton fichier d'entrée
    fichier_sortie = 'fichier_sortie.xlsx'  # à remplacer par ton fichier de sortie

    inverseur = InverseurNoms(fichier_entree, fichier_sortie)

    inverseur.traiter_fichier()


if __name__ == "__main__":
    main()
